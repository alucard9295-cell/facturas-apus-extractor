"""Genera un libro Excel descriptivo a partir de la base SQLite del extractor."""

from __future__ import annotations

import argparse
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

import db


OUTPUT_DEFAULT = Path(__file__).parent / "reportes" / "dashboard_financiero.xlsx"
NAVY = "111936"
CYAN = "72E6FF"
VIOLET = "A98CFF"
LIGHT = "E8EEFF"
MUTED = "6B7898"
GREEN = "DDF4C7"
RED = "FFD9E5"
THIN = Side(style="thin", color="D9E1F2")


def _rows(conn: sqlite3.Connection, query: str) -> list[dict[str, Any]]:
    return [dict(row) for row in conn.execute(query).fetchall()]


def _style_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + min(len(headers), 26))}{max(len(rows) + 1, 1)}"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=THIN)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                header = str(ws.cell(1, cell.column).value).lower()
                if any(word in header for word in ("total", "gasto", "iva", "precio", "valor", "promedio", "min", "max", "diferencia")):
                    cell.number_format = '$#,##0;[Red]-$#,##0'
                elif any(word in header for word in ("cantidad", "num_", "count", "id", "facturas", "items")):
                    cell.number_format = '#,##0.##'

    for index, header in enumerate(headers, start=1):
        values = [str(header)] + [str(row.get(header, "")) for row in rows[:100]]
        width = min(max(len(value) for value in values) + 2, 42)
        ws.column_dimensions[chr(64 + index)].width = max(width, 12)

    if rows:
        end = f"{chr(64 + len(headers))}{len(rows) + 1}"
        table = Table(displayName=f"Tabla{ws.title.replace(' ', '')}", ref=f"A1:{end}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)


def _write_data_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict[str, Any]]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    _style_sheet(ws, headers, rows)
    return ws


def _quality_rows(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return _rows(
        conn,
        """
        SELECT
            f.factura_id,
            f.numero_factura,
            f.proveedor_nombre,
            f.fecha_factura,
            f.subtotal,
            ROUND(COALESCE(SUM(fi.valor_total), 0), 2) AS suma_items,
            ROUND(COALESCE(SUM(fi.valor_total), 0) - COALESCE(f.subtotal, 0), 2) AS diferencia,
            CASE
                WHEN f.subtotal IS NULL THEN 'SIN_SUBTOTAL'
                WHEN ABS(COALESCE(SUM(fi.valor_total), 0) - f.subtotal) <= 5 THEN 'OK'
                ELSE 'REVISAR'
            END AS estado
        FROM facturas f
        LEFT JOIN factura_items fi ON fi.factura_id = f.factura_id
        GROUP BY f.factura_id
        ORDER BY f.fecha_factura, f.factura_id
        """,
    )


def _dashboard(wb: Workbook, metrics: dict[str, Any], monthly_ws, category_ws) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "DASHBOARD FINANCIERO · FACTURAS APUs"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} · Fuente: SQLite apu.db"
    ws["A2"].font = Font(italic=True, color=MUTED)

    cards = [
        ("A4", "Facturas", metrics["facturas"]),
        ("C4", "Total pagado", metrics["total_pagado"]),
        ("E4", "Proveedores", metrics["proveedores"]),
        ("G4", "Alertas calidad", metrics["alertas"]),
    ]
    for cell, label, value in cards:
        col = cell[0]
        ws[f"{col}4"] = label.upper()
        ws[f"{col}4"].font = Font(size=10, bold=True, color=MUTED)
        ws[f"{col}5"] = value
        ws[f"{col}5"].font = Font(size=18, bold=True, color=NAVY)
        ws[f"{col}5"].fill = PatternFill("solid", fgColor=LIGHT if label != "Alertas calidad" else RED)
        ws[f"{col}5"].number_format = '$#,##0;[Red]-$#,##0' if label == "Total pagado" else '#,##0.##'
        ws.merge_cells(f"{col}5:{chr(ord(col) + 1)}5")
        for row in range(4, 6):
            for column in range(ord(col) - 64, ord(col) - 62):
                ws.cell(row, column).border = Border(bottom=THIN)

    ws["A7"] = "Lectura rápida"
    ws["A7"].font = Font(size=13, bold=True, color=NAVY)
    ws["A8"] = "El dashboard separa tendencia mensual, concentración por proveedor, composición por categoría y calidad de extracción. Usa las hojas detalladas para auditar cada cifra."
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A8:H9")

    if monthly_ws.max_row > 1:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Total pagado por mes"
        chart.y_axis.title = "COP"
        chart.x_axis.title = "Mes"
        chart.add_data(Reference(monthly_ws, min_col=5, min_row=1, max_row=monthly_ws.max_row), titles_from_data=True)
        chart.set_categories(Reference(monthly_ws, min_col=1, min_row=2, max_row=monthly_ws.max_row))
        chart.height = 7
        chart.width = 13
        ws.add_chart(chart, "A11")

    if category_ws.max_row > 1:
        chart = PieChart()
        chart.title = "Gasto por categoría"
        chart.add_data(Reference(category_ws, min_col=2, min_row=1, max_row=category_ws.max_row), titles_from_data=True)
        chart.set_categories(Reference(category_ws, min_col=1, min_row=2, max_row=category_ws.max_row))
        chart.height = 7
        chart.width = 10
        ws.add_chart(chart, "J11")

    for column in "ABCDEFGH":
        ws.column_dimensions[column].width = 18


def generar_excel(db_path: str | Path | None = None, output: str | Path | None = None) -> Path:
    """Exporta todas las aristas financieras a un libro Excel independiente."""
    source = Path(db_path) if db_path else db.DB_PATH
    destination = Path(output) if output else OUTPUT_DEFAULT
    destination.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(source)
    conn.row_factory = sqlite3.Row
    try:
        facturas = _rows(conn, "SELECT * FROM facturas ORDER BY fecha_factura, factura_id")
        items = _rows(
            conn,
            """
            SELECT fi.item_id, fi.factura_id, f.numero_factura, f.fecha_factura,
                   f.proveedor_nombre, f.proyecto, fi.codigo_proveedor,
                   fi.descripcion_cruda, fi.unidad_medida, fi.cantidad,
                   fi.valor_unitario, fi.valor_total, fi.descuento,
                   im.nombre_normalizado, im.categoria
            FROM factura_items fi
            JOIN facturas f ON f.factura_id = fi.factura_id
            LEFT JOIN insumos_maestros im ON im.insumo_id = fi.insumo_id
            ORDER BY f.fecha_factura, fi.item_id
            """,
        )
        monthly = _rows(conn, "SELECT mes, num_facturas, gasto_materiales, iva_total, total_pagado FROM v_gasto_mensual ORDER BY mes")
        monthly_category = _rows(conn, "SELECT mes, categoria, gasto FROM v_gasto_mensual_categoria ORDER BY mes, categoria")
        categories = _rows(
            conn,
            """
            SELECT im.categoria, COUNT(fi.item_id) AS num_items,
                   ROUND(SUM(fi.valor_total), 2) AS gasto
            FROM factura_items fi
            JOIN insumos_maestros im ON im.insumo_id = fi.insumo_id
            GROUP BY im.categoria ORDER BY gasto DESC
            """,
        )
        providers = _rows(
            conn,
            """
            SELECT proveedor_nombre, proveedor_nit, COUNT(*) AS num_facturas,
                   ROUND(SUM(total_pagar), 2) AS total_pagado,
                   ROUND(AVG(total_pagar), 2) AS promedio_factura,
                   MAX(fecha_factura) AS ultima_factura
            FROM facturas GROUP BY proveedor_nombre, proveedor_nit ORDER BY total_pagado DESC
            """,
        )
        supplies = _rows(conn, "SELECT * FROM insumos_maestros ORDER BY categoria, nombre_normalizado")
        prices = _rows(conn, "SELECT * FROM v_insumo_precio_historico ORDER BY proveedor_nombre, insumo_id")
        quality = _quality_rows(conn)
    finally:
        conn.close()

    metrics = {
        "facturas": len(facturas),
        "total_pagado": sum((row.get("total_pagar") or 0) for row in facturas),
        "proveedores": len(providers),
        "alertas": sum(row["estado"] == "REVISAR" for row in quality),
    }

    wb = Workbook()
    wb.remove(wb.active)
    monthly_ws = _write_data_sheet(wb, "Mensual", ["mes", "num_facturas", "gasto_materiales", "iva_total", "total_pagado"], monthly)
    category_ws = _write_data_sheet(wb, "Categorias", ["categoria", "num_items", "gasto"], categories)
    _dashboard(wb, metrics, monthly_ws, category_ws)
    _write_data_sheet(wb, "Facturas", list(facturas[0].keys()) if facturas else ["factura_id", "proveedor_nombre", "fecha_factura", "total_pagar"], facturas)
    _write_data_sheet(wb, "Items", list(items[0].keys()) if items else ["item_id", "factura_id", "descripcion_cruda", "valor_total", "categoria"], items)
    _write_data_sheet(wb, "Mensual_Cat", ["mes", "categoria", "gasto"], monthly_category)
    _write_data_sheet(wb, "Proveedores", ["proveedor_nombre", "proveedor_nit", "num_facturas", "total_pagado", "promedio_factura", "ultima_factura"], providers)
    _write_data_sheet(wb, "Insumos", list(supplies[0].keys()) if supplies else ["insumo_id", "nombre_normalizado", "categoria", "unidad_estandar"], supplies)
    _write_data_sheet(wb, "Precios", list(prices[0].keys()) if prices else ["insumo_id", "proveedor_nombre", "precio_promedio", "precio_min", "precio_max"], prices)
    quality_ws = _write_data_sheet(wb, "Calidad", ["factura_id", "numero_factura", "proveedor_nombre", "fecha_factura", "subtotal", "suma_items", "diferencia", "estado"], quality)
    for row in quality_ws.iter_rows(min_row=2):
        state = row[7].value
        fill = GREEN if state == "OK" else RED if state == "REVISAR" else PatternFill(fill_type=None)
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill) if isinstance(fill, str) else fill

    wb.save(destination)
    return destination


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera dashboard financiero Excel desde apu.db")
    parser.add_argument("--db", default=str(db.DB_PATH))
    parser.add_argument("--salida", default=str(OUTPUT_DEFAULT))
    args = parser.parse_args()
    print(generar_excel(args.db, args.salida))


if __name__ == "__main__":
    main()
